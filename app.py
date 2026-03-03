import streamlit as st
import requests
import pandas as pd
from io import BytesIO
from base64 import b64encode
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

JIRA_EMAIL = st.secrets["jiraemail"]
JIRA_API = st.secrets["jiraapi"]
auth = b64encode(f'{JIRA_EMAIL}:{JIRA_API}'.encode()).decode()
JIRA_BASE = 'https://tc-isc.atlassian.net'
JIRA_H = {'Authorization': f'Basic {auth}', 'Accept': 'application/json'}

HARVEST_TOKEN = st.secrets["harvesttoken"]

HARVEST_ACCOUNT_ID = st.secrets["harvestid"]
HARVEST_H = {
    'Authorization': f'Bearer {HARVEST_TOKEN}',
    'Harvest-Account-Id': HARVEST_ACCOUNT_ID,
    'User-Agent': 'ConciliationReport'
}


JIRA_USERS = {
    'Karl Dionne' : '557058:91dd957f-3996-41de-9ecc-061691ca6316',
    'Thomas'      : '5c8911f853f3d02a237ee4c1',
    'Quincey'     : '557058:ec1448fb-dac8-4a2c-89b6-4551f2084e45',
    'Jordi'       : '557058:3de6efa9-3acf-49be-ad08-05ff2db1bcea',
    'Joseph'      : '5c6c1357ef824a130638a9c1',
    'Jason'       : '63092d16396fb6dc461b7d6a',
    'Lori'        : '557058:60ddf069-bc0e-4c31-a671-349be9c47ebd',
    'Luis'        : '557058:07eef689-7f5c-478b-9856-39211b3c2dc9',
    'Naveed'      : '557058:23bd8721-6ba0-44f5-8a92-bb1de64ab81f',
    'Valerii'     : '712020:12b2b91c-1e60-4cb7-9ac3-d750b4dee9b4',
    'Galina'      : '712020:ac75aa73-336f-4afb-99e0-bc4db35073e4',
    'Masoud'      : '712020:8b204003-6097-49ac-97d8-def4c99194d6',
    'Alex'        : '712020:e46ff762-af3d-4a46-9d35-20c75c062935',
    'Suresh'      : '557058:c47c07ad-7df2-4b02-a45d-ba7337776dfb',
    'Golam'       : '712020:82fe27e7-5ca3-48bf-b5ff-b8c85ec29594',
    'Amzad'       : '712020:9aecc43c-091e-463b-a3c4-269e5fc737d1',
    'Paul'        : '712020:69fa8cda-f9bc-45a9-ae77-59235682f1b2',
    'Neil'        : '712020:f1078d59-8813-471c-8324-e3c4ed0fecaf',
    'George'      : '712020:ffed9c51-c76c-472b-b816-93f8a420ad9d',
    'Apurv'       : '712020:308d4dd2-d126-44ec-bab8-09fac515367c',
    'Shawn'       : '712020:4620e38a-73dd-4164-a18c-db84e60c84fb',
    'Alessandro'  : '557058:c1a18f7d-2725-447b-a26a-f83823f60988',
    'Travis'      : '712020:dec097c6-a0d2-4e5c-a9f6-5fb83f6027e7',
    'Andy'        : '712020:3a275828-732a-40e4-9609-f64a4f51e53a',
    'Kenneth'     : '712020:aa3e0dd7-822e-40f6-bf81-4b69f86e056e',
}

HARVEST_USERS = {
    'Karl Dionne'  : 2816653,
    'Thomas'       : 3187249,
    'Quincey'      : 2983925,
    'Jordi'        : 3256061,
    'Joseph'       : 2983916,
    'Jason'        : 4353144,
    'Lori'         : 5298379,
    'Luis'         : 5324510,
    'Naveed'       : 2983923,
    'Valerii'      : 5385663,
    'Galina'       : 5334632,
    'Masoud'       : 5358508,
    'Alex'         : 4966821,
    'Suresh'       : 4169821,
    'Golam'        : 2983921,
    'Amzad'        : 4472620,
    'Paul'         : 5334363,
    'Neil'         : 4895224,
    'George'       : 4964592,
    'Apurv'        : 5460852,
    'Shawn'        : 5460854,
    'Alessandro'   : 5334630,
    'Travis'       : 5438373,
    'Andy'         : 5024635,
    'Kenneth'      : 5345474,
}

SDK_PROJECTS          = ['DFAPP']
PROMOBUILDER_PROJECTS = ['TISPT']
RADDAR_MAINT_PROJECTS = ['DFAGG', 'RASUP']
RADDAR2_PROJECTS      = ['RADD2']
OMNI_PROJECTS         = ['TCOMN']
DEDICATED_PROJECTS    = SDK_PROJECTS + PROMOBUILDER_PROJECTS + RADDAR_MAINT_PROJECTS + RADDAR2_PROJECTS + OMNI_PROJECTS

OUTPUT_GROUPS = [
    {
        'name'         : 'ADF FE',
        'jira_include' : None,
        'jira_exclude' : DEDICATED_PROJECTS,
        'harvest_id'   : 33523613,
        'users'        : ['Thomas', 'Quincey', 'Jordi', 'Joseph', 'Jason', 'Lori', 'Luis', 'Karl Dionne'],
    },
    {
        'name'         : 'ADF BE',
        'jira_include' : None,
        'jira_exclude' : DEDICATED_PROJECTS,
        'harvest_id'   : 33523613,
        'users'        : ['Naveed', 'Valerii', 'Galina', 'Masoud'],
    },
    {
        'name'         : 'SDK',
        'jira_include' : SDK_PROJECTS,
        'jira_exclude' : None,
        'harvest_id'   : 31344677,
        'users'        : ['Alex', 'Suresh', 'Lori', 'Jason', 'Joseph', 'Karl Dionne'],
    },
    {
        'name'         : 'PromoBuilder',
        'jira_include' : PROMOBUILDER_PROJECTS,
        'jira_exclude' : None,
        'harvest_id'   : 44631140,
        'users'        : ['Joseph', 'Thomas'],
    },
    {
        'name'         : 'Raddar Maintenance',
        'jira_include' : RADDAR_MAINT_PROJECTS,
        'jira_exclude' : None,
        'harvest_id'   : 37441461,
        'users'        : ['Galina', 'Luis', 'Jordi', 'Golam', 'Amzad', 'Paul', 'Masoud', 'Karl Dionne', 'Neil', 'George'],
    },
    {
        'name'         : 'Raddar2',
        'jira_include' : RADDAR2_PROJECTS,
        'jira_exclude' : None,
        'harvest_id'   : 44684187,
        'users'        : ['Masoud', 'Galina', 'Paul', 'Apurv', 'Shawn', 'Alessandro', 'Travis', 'Andy', 'Luis', 'Neil', 'Karl Dionne'],
    },
    {
        'name'         : 'OMNI',
        'jira_include' : OMNI_PROJECTS,
        'jira_exclude' : None,
        'harvest_id'   : 44566988,
        'users'        : ['Kenneth', 'Karl Dionne', 'Andy', 'Joseph', 'Golam', 'Amzad', 'Luis'],
    },
]

def secs_to_h(s):
    h, m = int(s // 3600), int((s % 3600) // 60)
    if h == 0: return round(m / 60, 2)
    return round(h + m / 60, 2)

def fetch_jira_issues(account_id, date_from, date_to, include=None, exclude=None):
    conditions = [
        f'worklogAuthor = "{account_id}"',
        f'worklogDate >= "{date_from}"',
        f'worklogDate <= "{date_to}"',
    ]
    if include:
        conditions.append(f'project IN ({", ".join(f'"{k}"' for k in include)})')
    if exclude:
        conditions.append(f'project NOT IN ({", ".join(f'"{k}"' for k in exclude)})')
    jql = ' AND '.join(conditions)
    issues, start = [], 0
    while True:
        r = requests.get(f'{JIRA_BASE}/rest/api/3/search', headers=JIRA_H,  #
                         params={'jql': jql, 'fields': 'worklog', 'maxResults': 50, 'startAt': start})
        if r.status_code != 200: break
        batch = r.json().get('issues', [])
        issues.extend(batch)
        if len(batch) < 50: break
        start += 50
    return issues

def fetch_jira_worklogs(issue_key):
    wl, start = [], 0
    while True:
        r = requests.get(f'{JIRA_BASE}/rest/api/3/issue/{issue_key}/worklog',
                         headers=JIRA_H, params={'maxResults': 100, 'startAt': start})
        if r.status_code != 200: break
        data = r.json()
        batch = data.get('worklogs', [])
        wl.extend(batch)
        if start + len(batch) >= data.get('total', 0): break
        start += len(batch)
    return wl

def get_jira_hours(account_id, date_from, date_to, include=None, exclude=None):
    issues = fetch_jira_issues(account_id, date_from, date_to, include, exclude)
    total = 0
    for issue in issues:
        wl_f = issue['fields'].get('worklog', {})
        worklogs = fetch_jira_worklogs(issue['key']) \
            if wl_f.get('total', 0) > len(wl_f.get('worklogs', [])) \
            else wl_f.get('worklogs', [])
        for e in worklogs:
            if e.get('author', {}).get('accountId') != account_id: continue
            if not (date_from <= e['started'][:10] <= date_to): continue
            total += e['timeSpentSeconds']
    return secs_to_h(total)

def get_harvest_hours(project_id, user_id, date_from, date_to):
    url = 'https://api.harvestapp.com/v2/time_entries'
    params = {'project_id': project_id, 'user_id': user_id,
              'from': date_from, 'to': date_to, 'per_page': 200}
    total, page = 0.0, 1
    while True:
        params['page'] = page
        r = requests.get(url, headers=HARVEST_H, params=params)
        if r.status_code != 200: return None
        data = r.json()
        for e in data.get('time_entries', []):
            total += e.get('hours', 0)
        if data.get('next_page') is None: break
        page = data['next_page']
    return round(total, 2)

# Streamlit UI
st.title("KPDI Logged Hours Report")

default_from = datetime(2026, 2, 1)
default_to = datetime(2026, 2, 27)

date_from = st.date_input("Date From", value=default_from)
date_to = st.date_input("Date To", value=default_to)

DATE_FROM = date_from.strftime("%Y-%m-%d")
DATE_TO = date_to.strftime("%Y-%m-%d")

if st.button("Generate Report"):
    with st.spinner("Fetching data..."):
        results = {}
        for group in OUTPUT_GROUPS:
            gname = group['name']
            results[gname] = {}
            st.write(f'--- {gname} ---')
            for uname in group['users']:
                jira_h = 0.0
                harvest_h = 0.0

                if uname in JIRA_USERS:
                    jira_h = get_jira_hours(
                        JIRA_USERS[uname], DATE_FROM, DATE_TO,
                        group['jira_include'], group['jira_exclude']
                    )

                if uname in HARVEST_USERS:
                    h = get_harvest_hours(group['harvest_id'], HARVEST_USERS[uname], DATE_FROM, DATE_TO)
                    if h is not None:
                        harvest_h = h

                results[gname][uname] = {'jira': jira_h, 'harvest': harvest_h}
                diff = round(jira_h - harvest_h, 2)
                st.write(f'  {uname:<25} Jira: {jira_h:>6}h  Harvest: {harvest_h:>6}h  Diff: {diff:>+.2f}h')

        # Build DataFrame
        rows = []
        for group in OUTPUT_GROUPS:
            gname = group['name']
            group_data = results[gname]

            jira_total = 0.0
            harvest_total = 0.0

            rows.append({
                'Project - User ': gname,
                'Harvest (h)': '',
                'Jira (h)': '',
                'Diff (h)': '',
            })

            for uname in group['users']:
                d = group_data.get(uname, {'jira': 0, 'harvest': 0})
                j = d['jira']
                h = d['harvest']
                diff = round(j - h, 2)
                jira_total += j
                harvest_total += h
                rows.append({
                    'Project - User ': f'    {uname}',
                    'Harvest (h)': h,
                    'Jira (h)': j,
                    'Diff (h)': diff,
                })

            total_diff = round(jira_total - harvest_total, 2)
            rows.append({
                'Project - User ': '    TOTAL',
                'Harvest (h)': round(harvest_total, 2),
                'Jira (h)': round(jira_total, 2),
                'Diff (h)': total_diff,
            })
            rows.append({'Project - User ': '', 'Harvest (h)': '', 'Jira (h)': '', 'Diff (h)': ''})

        df = pd.DataFrame(rows)

       
        st.dataframe(df)

        #  Excel 
        excel_buffer = BytesIO()
        with pd.ExcelWriter(excel_buffer, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='JiraHarvest')
            ws = writer.sheets['Conciliacion']

            header_fill = PatternFill('solid', fgColor='1F4E79')
            group_fill = PatternFill('solid', fgColor='2E75B6')
            total_fill = PatternFill('solid', fgColor='D6E4F0')
            red_font = Font(color='C00000', bold=True)
            white_bold = Font(color='FFFFFF', bold=True)
            bold = Font(bold=True)

            for cell in ws[1]:
                cell.font = white_bold
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

            for row in ws.iter_rows(min_row=2):
                label = str(row[0].value or '')
                if label and not label.startswith('    '):
                    for cell in row:
                        cell.fill = group_fill
                        cell.font = white_bold
                elif label.strip() == 'TOTAL':
                    for cell in row:
                        cell.fill = total_fill
                        cell.font = bold
                diff_cell = row[3]
                if isinstance(diff_cell.value, (int, float)) and diff_cell.value < 0:
                    diff_cell.font = red_font

            col_widths = [35, 14, 14, 16]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

        excel_buffer.seek(0)

        st.download_button(
            label="Download Excel Report",
            data=excel_buffer,
            file_name=f'MergedJiraHarvest_{DATE_FROM}_to_{DATE_TO}.xlsx',
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

    st.success("Report generated!")
